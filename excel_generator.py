from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from questions import QUESTIONS

def create_excel_file(user_data):
    """
    Foydalanuvchi ma'lumotlaridan Excel fayl yaratish
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ariza"
    
    # Sarlavha
    ws['A1'] = 'ISHGA MUROJAAT QILUVCHI MA\'LUMOTLARI'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:B1')
    
    # Lavozim
    ws['A2'] = 'Lavozim:'
    ws['B2'] = user_data.get('position', '')
    ws['A2'].font = Font(bold=True)
    ws['A2'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    ws['A3'] = 'Ariza sanasi:'
    ws['B3'] = datetime.now().strftime('%d.%m.%Y')
    ws['A3'].font = Font(bold=True)
    ws['A3'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    # Bo'sh qator
    row = 5
    
    for question in QUESTIONS:
        key = question['key']
        question_num = question['id']
        question_label = question.get('label', question['question'].split('\n')[0])
        
        if key in user_data:
            ws[f'A{row}'] = f'{question_num}. {question_label}'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            value = user_data[key]
            cell = ws[f'B{row}']
            
            if isinstance(value, str) and (value.startswith('http://') or value.startswith('https://')):
                # Telegram link bo'lsa, display text va hyperlink qo'shish
                if 't.me' in value:
                    cell.value = '🔗 Telegram xabar'
                    cell.hyperlink = value
                    cell.font = Font(color='0563C1', underline='single')
                else:
                    cell.value = value
                    cell.hyperlink = value
                    cell.font = Font(color='0563C1', underline='single')
            else:
                cell.value = value
            
            row += 1
    
    # Ustunlar kengligini sozlash
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 60
    
    # Faylni saqlash
    filename = f"ariza_{user_data.get('full_name', 'user').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join('temp', filename)
    
    # temp papkasini yaratish
    os.makedirs('temp', exist_ok=True)
    
    wb.save(filepath)
    return filepath
